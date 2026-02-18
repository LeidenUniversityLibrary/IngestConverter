import re
import sys
import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
import colorist
from colorama import Fore, Back, Style, init
init(convert=True)

usedir = os.getcwd() + "\\" 
ingestcsvfile = usedir + "ingest.csv"
errorCount = 0
errorMessage = ""
writeLine = "mms_id;permanent_call_number;representation_label;access_rights;file_path;file_label;collection_id\n"

print("╻ ╻┏━╸╻  ┏━╸┏━┓┏┳┓┏━╸   ╺┳╸┏━┓")
print("┃╻┃┣╸ ┃  ┃  ┃ ┃┃┃┃┣╸     ┃ ┃ ┃")
print("┗┻┛┗━╸┗━╸┗━╸┗━┛╹ ╹┗━╸    ╹ ┗━┛")
print(Fore.GREEN +" _____                           _           _____  _____  _____  _____ ")
print("/  __ \                         | |         |____ ||  _  ||  _  ||  _  |")
print("| /  \/ ___  _ ____   _____ _ __| |_ ___ _ __   / /| |/' || |/' || |/' |")
print("| |    / _ \| '_ \ \ / / _ \ '__| __/ _ \ '__|  \ \|  /| ||  /| ||  /| |")
print("| \__/\ (_) | | | \ V /  __/ |  | ||  __/ | .___/ /\ |_/ /\ |_/ /\ |_/ /")
print(" \____/\___/|_| |_|\_/ \___|_|   \__\___|_| \____/  \___/  \___/  \___/ ")
print("                                                                        " + Style.RESET_ALL)
                                                                        


fcount = 0
fdict = {}

for pathfile in os.listdir(usedir):
	if '.xlsx' in pathfile and not '~$' in pathfile:
		fcount+=1
		fdict[fcount] = pathfile

if fcount == 0:
	print("")
	print("No Excel files found to convert!")
	isExit = input("Close this window or type \"exit\" to close: ")
	if isExit == "exit":
		sys.exit(2)
elif fcount > 1:
	print("")
	print("Please choose a file to convert:")
	for file in fdict:
		print(str(file) + ": " + fdict[file])
	print("")
	choosevar = input("File: ")
else:
	choosevar = '1'
if int(choosevar) not in fdict:
	print("Please choose a number in the list")
	exit()

xlsxfile = usedir + fdict[int(choosevar)]
wb = load_workbook(xlsxfile)
ws = wb.active

for index, row in enumerate(ws.values):
	if index > 0:
		mmsid = ws.cell(row=index+1, column=1).value
		shelfmark = ws.cell(row=index+1, column=2).value
		replabel = ws.cell(row=index+1, column=3).value
		arights = ws.cell(row=index+1, column=4).value
		fpath = ws.cell(row=index+1, column=6).value
		flabel = ws.cell(row=index+1, column=7).value
		colid = ws.cell(row=index+1, column=8).value
		
		#Check if mmsid valid
		if type(mmsid) != type('str'):
			errorCount += 1
			errorMessage = errorMessage + Fore.RED + "ERROR " + str(errorCount) + Style.RESET_ALL + ": MMSID on line " + str(index+1) + " is not a valid MMSID! \n"
			errorMessage = errorMessage + str(mmsid) + " is an integer, but should be a string! \n\n"
		else:
			if mmsid[0] != '9' or mmsid[1] != '9' or mmsid[len(mmsid)-4] != '2' or mmsid[len(mmsid)-3] != '7' or mmsid[len(mmsid)-2] != '1' or mmsid[len(mmsid)-1] != '1':
				errorCount+=1
				errorMessage = errorMessage + Fore.RED + "ERROR " + str(errorCount) + Style.RESET_ALL + ": MMSID on line " + str(index+1) + " is not a valid MMSID! \n"
				errorMessage = errorMessage + (mmsid) + " does not start with 99 or end with 2711! \n\n"
				
		
		#Check if all fields are filled
		for i, cell in enumerate(row):
			if cell == None:
				errorCount+=1
				errorMessage = errorMessage + Fore.RED + "ERROR " + str(errorCount)  + Style.RESET_ALL + ": Line " + str(index+1) + " has empty values! \n"
				errorMessage = errorMessage + "Is missing: \"" + ws.cell(row=1, column=i).value + "\" in column " + str(i) + "  \n\n"
		
		if errorCount == 0:
			writeLine = writeLine + (mmsid + ";" + shelfmark + ";" + replabel + ";" + arights + ";" + fpath + ";" + flabel + ";" + colid + "\n")

if errorCount > 0:
	print(Fore.RED + Style.BRIGHT + "WARNING! " + Style.RESET_ALL + str(errorCount) + " ERRORS HAVE BEEN FOUND! \n")
	print(errorMessage)
else:
	print(Fore.GREEN + "Success! " + Style.RESET_ALL + "CSV file has been created")
	with open(ingestcsvfile, "w", encoding="utf-8-sig") as a:
		a.write(writeLine)

isExit = ""
while isExit != "exit":
	isExit = input("Close this window or type \"exit\" to close: ")
	if isExit == "exit":
		sys.exit(2)
